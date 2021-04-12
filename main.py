import getProductList
import openpyxl

def getTable() :
    excelFile = openpyxl.load_workbook('Book1.xlsx')

    # sheet1.title = 'Sheet1'
    # B1 = sheet[B1]
    # sheet.cell(row=1, column=1).value = 10
    # B1.value
    return excelFile

# Press the green button in the gutter to run the script.
if __name__ == '__main__':

    excelFile = getTable()

    sheet = excelFile.active  # 엑셀은 최소한 1개이상의 sheet는 존재. 기본적으로 존재하는 sheet을 가져올때 이렇게.

    # print(sheet.cell(row=175, column=11).value == None)


    #Param set
    current_row = 2
    next_row = True


    while(next_row) :
        # Get URL
        # 테이블 명 : 4     메뉴얼 링크 ?    	From 상위 15	        To 하위 16	    설명 17
        cellValue = sheet.cell(row=current_row, column=4).value
        aa = getProductList.getAPI(f"http://isparkmom/apriso/Help/en-us/DB/xmls/Table_{cellValue}.xml")

        if "Error Code" != aa :

            sheet.cell(row=current_row, column=15).value = aa['FK_From']
            sheet.cell(row=current_row, column=16).value = aa['FK_To']
            sheet.cell(row=current_row, column=17).value = aa['SDESCR']

            print(current_row ,aa['FK_From'], aa['FK_To'], aa['SDESCR'] )

        # Row ++
        current_row += 1
        # While Condition
        # next_row = False
        if( sheet.cell(row=current_row, column=4).value == None) :
            next_row = False

    excelFile.save('fommater_new.xlsx')

    #
    # # 8 , 9
    # print( aa['FK_From'], aa['FK_To'] )

